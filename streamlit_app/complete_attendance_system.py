import streamlit as st
import os
import json
import pandas as pd
import numpy as np
from datetime import datetime
import time
import cv2
from PIL import Image
import io
import base64
import threading
import queue
import plotly.express as px
import sys
import joblib
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# Set page config
st.set_page_config(
    page_title="MSN Global IT - AI Attendance System",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Initialize session state
if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False
if 'username' not in st.session_state:
    st.session_state.username = None
if 'auto_monitoring' not in st.session_state:
    st.session_state.auto_monitoring = True
if 'auto_refresh' not in st.session_state:
    st.session_state.auto_refresh = True
if 'camera_active' not in st.session_state:
    st.session_state.camera_active = False
if 'face_detection_active' not in st.session_state:
    st.session_state.face_detection_active = False
if 'refresh_timestamp' not in st.session_state:
    st.session_state.refresh_timestamp = 0
if 'last_face_check' not in st.session_state:
    st.session_state.last_face_check = 0
if 'uploaded_csv_data' not in st.session_state:
    st.session_state.uploaded_csv_data = None
if 'uploaded_csv_filename' not in st.session_state:
    st.session_state.uploaded_csv_filename = None
if 'last_notification' not in st.session_state:
    st.session_state.last_notification = None
if 'notification_type' not in st.session_state:
    st.session_state.notification_type = None


# Custom CSS for black, red, white theme
st.markdown("""
<style>
    .main {
        background-color: #000000;
        color: white;
    }
    .stButton > button {
        background-color: #DC143C;
        color: white;
        border: 2px solid #DC143C;
        border-radius: 10px;
        padding: 10px 20px;
        font-weight: bold;
        min-height: 50px;
        width: 100%;
        text-align: center;
    }
    .stButton > button:hover {
        background-color: #B22222;
        border-color: #B22222;
    }
    .stButton > button:active {
        background-color: #8B0000;
        border-color: #8B0000;
    }
    /* Ensure equal width for quick control buttons */
    .stButton > button[data-testid="gen_report"],
    .stButton > button[data-testid="monthly_report"],
    .stButton > button[data-testid="analytics"],
    .stButton > button[data-testid="clear_entries"],
    .stButton > button[data-testid="export_data"] {
        width: 100% !important;
        min-width: 120px !important;
        max-width: 200px !important;
        margin: 0 auto !important;
        display: block !important;
    }
    .stSelectbox > div > div > div {
        background-color: #000000;
        color: white;
        border: 2px solid #DC143C;
    }
    .stTextInput > div > div > input {
        background-color: #000000;
        color: white;
        border: 2px solid #DC143C;
    }
    .stDataFrame {
        background-color: #000000;
        color: white;
    }
    .stMetric {
        background-color: #000000;
        color: white;
    }
    .stMarkdown {
        color: white;
    }
    .stExpander {
        background-color: #000000;
        color: white;
        border: 2px solid #DC143C;
    }
    .stSidebar {
        background-color: #000000;
        color: white;
    }
    .stSidebar .stMarkdown {
        color: white;
    }
    .stSidebar .stButton > button {
        background-color: #DC143C;
        color: white;
        border: 2px solid #DC143C;
    }
    .stSidebar .stButton > button:hover {
        background-color: #B22222;
        border-color: #B22222;
    }
    .stSidebar .stButton > button:active {
        background-color: #8B0000;
        border-color: #8B0000;
    }
    .stTabs [data-baseweb="tab-list"] {
        background-color: #000000;
        color: white;
    }
    .stTabs [data-baseweb="tab"] {
        background-color: #000000;
        color: white;
    }
    .stAlert {
        background-color: #000000;
        color: white;
        border: 2px solid #DC143C;
    }
    .stSuccess {
        background-color: #000000;
        color: #4CAF50;
        border: 2px solid #4CAF50;
    }
    .stError {
        background-color: #000000;
        color: #f44336;
        border: 2px solid #f44336;
    }
    .stWarning {
        background-color: #000000;
        color: #ff9800;
        border: 2px solid #ff9800;
    }
    .stInfo {
        background-color: #000000;
        color: #2196F3;
        border: 2px solid #2196F3;
    }
</style>
""", unsafe_allow_html=True)

# Paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_DIR = os.path.join(BASE_DIR, 'excels')
os.makedirs(EXCEL_DIR, exist_ok=True)

# Default credentials - Only manager
DEFAULT_ADMIN = {
    "msnglobalit": "msnglobalit123"
}

# Load credentials
def load_credentials():
    cred_file = "credentials.json"
    if os.path.exists(cred_file):
        try:
            with open(cred_file, 'r') as f:
                return json.load(f)
        except:
            return DEFAULT_ADMIN
    return DEFAULT_ADMIN

# Save credentials
def save_credentials(credentials):
    with open("credentials.json", 'w') as f:
        json.dump(credentials, f, indent=2)

# Authenticate user
def authenticate(username, password):
    credentials = load_credentials()
    return username in credentials and credentials[username] == password

# Face recognition integration
RECOGNIZED_ID_FILE = os.path.join(os.path.dirname(BASE_DIR), 'face_recognition_app', 'shared', 'recognized_id.json')

def check_face_recognition():
    """Check for new face recognition from JSON file - Auto-save detected IDs"""
    try:
        if os.path.exists(RECOGNIZED_ID_FILE):
            if os.path.getsize(RECOGNIZED_ID_FILE) > 0:
                with open(RECOGNIZED_ID_FILE, 'r') as f:
                    data = json.load(f)
                
                emp_id = data.get('employee_id')
                name = data.get('name')
                unique_id = data.get('unique_id')
                
                if emp_id and unique_id:
                    # Always mark attendance (allow multiple entries per day)
                    if not name:
                        # If name is missing, try to get it from employee data
                        try:
                            employee_df = load_employee_data()
                            if employee_df is not None and not employee_df.empty:
                                emp_row = employee_df[employee_df['Employee ID'] == emp_id]
                                if not emp_row.empty:
                                    name = emp_row['Name'].iloc[0]
                        except Exception as e:
                            print(f"Error getting employee name: {e}")
                            name = f"Employee {emp_id}"
                    
                    success, message = mark_attendance(emp_id, name or f"Employee {emp_id}")
                    if success:
                        # Clear the JSON file after processing
                        try:
                            os.remove(RECOGNIZED_ID_FILE)
                        except:
                            pass
                        return True, f"✅ Face Recognition: {name or emp_id} ({emp_id}) detected and marked!"
                    else:
                        return False, f"❌ Face Recognition: {message}"
                else:
                    return False, "❌ Face Recognition: Invalid data in JSON file"
            else:
                return False, "ℹ️ Face Recognition: No new detection data"
        else:
            return False, "ℹ️ Face Recognition: No detection file found"
    except json.JSONDecodeError:
        # Clear corrupted JSON file
        try:
            os.remove(RECOGNIZED_ID_FILE)
        except:
            pass
        return False, "❌ Face Recognition: Corrupted detection data"
    except Exception as e:
        print(f"Error reading face recognition JSON: {e}")
        return False, f"❌ Face Recognition: Error reading data"
    
    return False, None

def get_face_recognition_status():
    """Get current face recognition status with detailed information"""
    if os.path.exists(RECOGNIZED_ID_FILE):
        try:
            with open(RECOGNIZED_ID_FILE, 'r') as f:
                data = json.load(f)
                emp_id = data.get('employee_id', 'Unknown')
                name = data.get('name', 'Unknown')
                timestamp = data.get('timestamp', '')
                status = data.get('status', 'Unknown')
                
                if timestamp:
                    try:
                        # Parse timestamp and show time
                        dt = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                        time_str = dt.strftime('%H:%M:%S')
                        return f"🟢 Active - {name} ({emp_id}) at {time_str} - {status}"
                    except:
                        return f"🟢 Active - {name} ({emp_id}) - {status}"
                else:
                    return f"🟢 Active - {name} ({emp_id}) - {status}"
        except Exception as e:
            return "🟡 Connected - JSON file exists but has errors"
    else:
        return "🔴 Not Connected - Start face recognition app"

# Simulate face detection (since we can't use camera in Streamlit)
def simulate_face_detection():
    """Simulate face detection with actual employee IDs from CSV"""
    # Load actual employee data
    emp_file = os.path.join(os.path.dirname(BASE_DIR), 'face_recognition_app', 'shared', 'employees_data.csv')
    employees = []
    names = []
    
    try:
        if os.path.exists(emp_file):
            df = pd.read_csv(emp_file)
            employees = df['Employee ID'].tolist()
            names = df['Name'].tolist()
        else:
            # Fallback to default data
            employees = ["MSN001", "MSN002", "MSN003", "MSN004", "MSN005"]
            names = ["Ramsha Tariq", "Tehreem Siddiqui", "Rayyan Ahmad", "Maryam Sheikh", "Samreen Fatima"]
    except:
        # Fallback to default data
        employees = ["MSN001", "MSN002", "MSN003", "MSN004", "MSN005"]
        names = ["Ramsha Tariq", "Tehreem Siddiqui", "Rayyan Ahmad", "Maryam Sheikh", "Samreen Fatima"]
    
    # Simulate detection with actual employee IDs
    if np.random.random() < 0.4:  # 40% chance of detection
        # Select a random employee from the actual list
        idx = np.random.randint(0, len(employees))
        emp_id = employees[idx]
        name = names[idx]
        return {
            "employee_id": emp_id,
            "name": name,
            "timestamp": datetime.now().isoformat(),
            "status": "recognized",
            "unique_id": f"{emp_id}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}"
        }
    return None

# Load employee data
def load_employee_data():
    # First check if we have uploaded CSV data in session state
    if st.session_state.uploaded_csv_data is not None:
        return pd.DataFrame(st.session_state.uploaded_csv_data)
    
    # Otherwise load from file
    emp_file = os.path.join(os.path.dirname(BASE_DIR), 'face_recognition_app', 'shared', 'employees_data.csv')
    if os.path.exists(emp_file):
        try:
            return pd.read_csv(emp_file)
        except:
            pass
    
    # Create default employee data with actual names
    default_data = {
        'Employee ID': ['MSN001', 'MSN002', 'MSN003', 'MSN004', 'MSN005', 'MSN006', 'MSN007', 'MSN009'],
        'Name': ['Ramsha Tariq', 'Tehreem Siddiqui', 'Rayyan Ahmad', 'Maryam Sheikh', 'Samreen Fatima', 'Taskeen Abbas', 'Muhammad Shaf', 'Hammad Hassan']
    }
    return pd.DataFrame(default_data)

def generate_wide_excel(df, filename):
    """Generate wide format Excel report with proper totals and formatting"""
    # Check if dataframe is empty
    if df.empty:
        print("Warning: No data to generate Excel report")
        return None
    
    try:
        # Convert Date column to datetime
        df['Date'] = pd.to_datetime(df['Date'])
        
        # Format date display based on OS
        if os.name == 'nt':
            df['Date_Display'] = df['Date'].dt.strftime('%B %#d, %Y')
        else:
            df['Date_Display'] = df['Date'].dt.strftime('%B %-d, %Y')
        
        # Create pivot table
        pivot_df = df.pivot_table(
            index=['Employee ID', 'Name'],
            columns='Date_Display',
            values='Status',
            aggfunc='first'
        ).reset_index()
        
        # Fill NaN values with empty string
        pivot_df = pivot_df.fillna('')
        
        # Calculate totals correctly - only count status columns (skip Employee ID and Name)
        status_columns = [col for col in pivot_df.columns if col not in ['Employee ID', 'Name']]
        
        # Calculate totals for each employee
        if status_columns:
            pivot_df['Total Present'] = pivot_df[status_columns].apply(lambda x: (x == 'Present').sum(), axis=1)
            pivot_df['Total Late'] = pivot_df[status_columns].apply(lambda x: (x == 'Late').sum(), axis=1)
            pivot_df['Total Absent'] = pivot_df[status_columns].apply(lambda x: (x == 'Absent').sum(), axis=1)
            pivot_df['Total Leaves'] = pivot_df[status_columns].apply(lambda x: (x == 'Leave').sum(), axis=1)
        else:
            # If no status columns, add empty total columns
            pivot_df['Total Present'] = 0
            pivot_df['Total Late'] = 0
            pivot_df['Total Absent'] = 0
            pivot_df['Total Leaves'] = 0
        
        # Generate filename
        actual_filename = f'Attendance_Report_{get_month_year()}.xlsx'
        
        # Create styled Excel report
        try:
            create_styled_excel_report(pivot_df, actual_filename)
            return actual_filename
        except Exception as e:
            print(f"Error creating styled Excel report: {e}")
            return None
            
    except Exception as e:
        print(f"Error in generate_wide_excel: {e}")
        return None

# Mark attendance
def mark_attendance(emp_id, name):
    """Mark attendance for an employee with proper Excel format"""
    now = datetime.now()
    today_date_col = get_today_date()
    entry_time = now.replace(second=0, microsecond=0).time()
    status = fallback_time_based_prediction(entry_time)
    
    print(f"⏰ Time-based prediction for {emp_id}: {entry_time.strftime('%H:%M')} - Status: {status}")
    
    raw_file = os.path.join(EXCEL_DIR, f'Attendance_Raw_{get_month_year()}.xlsx')
    
    try:
        if os.path.exists(raw_file):
            df_existing = pd.read_excel(raw_file)
        else:
            df_existing = pd.DataFrame(columns=['Employee ID', 'Name', 'Date', 'Entry_Time', 'Status'])
    except Exception as e:
        print(f"Error loading existing data: {e}")
        df_existing = pd.DataFrame(columns=['Employee ID', 'Name', 'Date', 'Entry_Time', 'Status'])

    # Check for duplicate entry
    already_marked = False
    if not df_existing.empty:
        today_entries = df_existing[df_existing['Date'] == today_date_col]
        if not today_entries.empty:
            already_marked = (today_entries['Employee ID'] == emp_id).any()

    if already_marked:
        return False, f"Employee {name} ({emp_id}) already marked attendance today"

    # Create new record
    new_record = pd.DataFrame([{
        'Employee ID': emp_id,
        'Name': name,
        'Date': today_date_col,
        'Entry_Time': entry_time,
        'Status': status
    }])

    # Save combined data
    try:
        df_combined = pd.concat([df_existing, new_record], ignore_index=True)
        df_combined.to_excel(raw_file, index=False)
        print(f"✅ Saved attendance data for {emp_id}")
    except Exception as e:
        print(f"❌ Error saving data: {e}")
        return False, f"Failed to save attendance data for {name} ({emp_id})"
    
    # Auto-generate styled Excel report with wide format
    try:
        actual_filename = generate_wide_excel(df_combined, f'Attendance_Report_{get_month_year()}.xlsx')
        if actual_filename:
            st.cache_data.clear()
            return True, f"Marked {name} ({emp_id}) as {status} at {entry_time.strftime('%H:%M')}"
        else:
            print(f"Warning: Excel generation failed")
            return True, f"Marked {name} ({emp_id}) as {status} at {entry_time.strftime('%H:%M')}"
    except Exception as e:
        print(f"Warning: Excel styling failed: {e}")
        return True, f"Marked {name} ({emp_id}) as {status} at {entry_time.strftime('%H:%M')}"
    
    return False, f"Failed to mark attendance for {name} ({emp_id})"

def predict_attendance_with_rf_model(emp_id, name, entry_time):
    """Use the best RF model to predict attendance status"""
    try:
        # Load the best RF model
        model_path = os.path.join(os.path.dirname(BASE_DIR), 'face_recognition_app', 'shared', 'best_rf_model.pkl')
        
        if not os.path.exists(model_path):
            print(f"❌ RF model not found at: {model_path}")
            # Fallback to simple time-based logic
            return fallback_time_based_prediction(entry_time)
        
        # Load the model
        model = joblib.load(model_path)
        print(f"✅ RF model loaded successfully")
        
        # Create a sample DataFrame for prediction
        sample_data = pd.DataFrame([{
            'Employee ID': emp_id,
            'Name': name,
            'Date': get_today_date(),
            'Entry_Time': entry_time
        }])
        
        # Import the prediction function from the correct path
        # Use the modules from face_recognition_app
        face_recognition_modules_path = os.path.join(os.path.dirname(BASE_DIR), 'face_recognition_app', 'modules')
        sys.path.insert(0, face_recognition_modules_path)  # Insert at beginning to prioritize
        
        # Clear any cached imports to force reload
        import importlib
        for module_name in ['predict_attendance', 'data_cleaning', 'feature_engineering', 'leave_calculator']:
            if module_name in sys.modules:
                del sys.modules[module_name]
        
        from predict_attendance import predict_attendance
        
        # Use the RF model to predict
        try:
            df_predictions, leave_summary = predict_attendance(sample_data, model)
            
            # Get the predicted status
            predicted_status = df_predictions['Status'].iloc[0]
            print(f"🤖 RF Model predicted: {predicted_status}")
            
            return predicted_status
        except Exception as e:
            print(f"❌ RF Model prediction failed: {e}")
            # Fallback to simple time-based logic
            return fallback_time_based_prediction(entry_time)
        
    except Exception as e:
        print(f"❌ Error using RF model: {e}")
        # Fallback to simple time-based logic
        return fallback_time_based_prediction(entry_time)

def fallback_time_based_prediction(entry_time):
    """Fallback to simple time-based prediction if RF model fails"""
    # Define office hours and late threshold
    office_start_time = datetime.strptime("09:00", "%H:%M").time()  # 9:00 AM
    late_threshold = datetime.strptime("09:15", "%H:%M").time()     # 9:15 AM (15 minutes grace)
    
    # Determine attendance status based on time
    if entry_time <= office_start_time:
        return "Present"
    elif entry_time <= late_threshold:
        return "Present"  # Still considered present within grace period
    else:
        return "Late"

def mark_attendance_with_time(emp_id, name, time_str):
    """Mark attendance with a specific time for testing purposes"""
    # Parse the time string
    try:
        test_time = datetime.strptime(time_str, "%H:%M").time()
    except:
        return False, f"Invalid time format: {time_str}"
    
    # Get current date
    today_date_col = get_today_date()
    
    # Use RF model for prediction
    status = predict_attendance_with_rf_model(emp_id, name, test_time)
    
    print(f"⏰ RF Model test prediction for {emp_id}: {test_time.strftime('%H:%M')} - Status: {status}")
    
    # Load existing data
    raw_file = os.path.join(EXCEL_DIR, f'Attendance_Raw_{get_month_year()}.xlsx')
    
    try:
        if os.path.exists(raw_file):
            df_existing = pd.read_excel(raw_file)
        else:
            df_existing = pd.DataFrame(columns=['Employee ID', 'Name', 'Date', 'Entry_Time', 'Status'])
    except Exception as e:
        df_existing = pd.DataFrame(columns=['Employee ID', 'Name', 'Date', 'Entry_Time', 'Status'])

    # Check for duplicate
    already_marked = False
    if not df_existing.empty:
        today_entries = df_existing[df_existing['Date'] == today_date_col]
        if not today_entries.empty:
            already_marked = (today_entries['Employee ID'] == emp_id).any()

    if already_marked:
        return False, f"Employee {name} ({emp_id}) already marked attendance today"

    # Create new record with test time
    new_record = pd.DataFrame([{
        'Employee ID': emp_id,
        'Name': name,
        'Date': today_date_col,
        'Entry_Time': test_time,
        'Status': status
    }])

    # Save combined data
    df_combined = pd.concat([df_existing, new_record], ignore_index=True)
    df_combined.to_excel(raw_file, index=False)
    
    # Auto-generate styled Excel report with wide format
    try:
        actual_filename = generate_wide_excel(df_combined, f'Attendance_Report_{get_month_year()}.xlsx')
        if actual_filename:
            st.cache_data.clear()
            return True, f"Test: Marked {name} ({emp_id}) as {status} at {test_time.strftime('%H:%M')}"
        else:
            print(f"Warning: Excel generation failed")
            return True, f"Test: Marked {name} ({emp_id}) as {status} at {test_time.strftime('%H:%M')}"
    except Exception as e:
        print(f"Warning: Excel styling failed: {e}")
        return True, f"Test: Marked {name} ({emp_id}) as {status} at {test_time.strftime('%H:%M')}"
    
    return False, f"Failed to mark test attendance for {name} ({emp_id})"

def auto_update_daily_excel():
    """Auto-update daily Excel report in wide format with totals"""
    try:
        raw_file = os.path.join(EXCEL_DIR, f'Attendance_Raw_{get_month_year()}.xlsx')
        if os.path.exists(raw_file):
            df = pd.read_excel(raw_file)
            if not df.empty:
                # Generate wide format Excel report with totals
                actual_filename = generate_wide_excel(df, f'Attendance_Report_{get_month_year()}.xlsx')
                if actual_filename:
                    return True
                else:
                    print("Warning: Excel generation failed in auto_update")
                    return False
    except Exception as e:
        print(f"Error auto-updating Excel: {e}")
        return False

# Helper function for consistent date formatting
def get_today_date():
    """Get today's date in simplified format"""
    return datetime.now().strftime("%d/%m/%Y")

def get_today_date_display():
    """Get today's date for display purposes"""
    return datetime.now().strftime("%A, %d %B %Y")

def get_month_year():
    """Get month and year for file naming"""
    return datetime.now().strftime("%B_%Y")

import base64
import joblib
import sys
import os

# Add the face_recognition_app modules to the path
face_recognition_path = os.path.join(os.path.dirname(BASE_DIR), 'face_recognition_app')
if face_recognition_path not in sys.path:
    sys.path.append(face_recognition_path)

def get_logo_base64():
    """Get logo as base64 string"""
    try:
        logo_path = os.path.join(os.path.dirname(BASE_DIR), 'assets', 'logo.png')
        if os.path.exists(logo_path):
            with open(logo_path, "rb") as image_file:
                return base64.b64encode(image_file.read()).decode()
        else:
            # Return a simple placeholder if logo doesn't exist
            return ""
    except Exception as e:
        print(f"Error loading logo: {e}")
        return ""

# Excel styling function
def style_excel(file_path):
    """Apply professional styling to Excel file with black, red, white theme and total statistics"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        # Get the original data before inserting rows
        original_data = []
        for row in ws.iter_rows(min_row=2):  # Skip header row
            row_data = [cell.value for cell in row]
            if any(row_data):  # Only add non-empty rows
                original_data.append(row_data)

        # Calculate total statistics from original data
        total_present = 0
        total_late = 0
        total_absent = 0
        total_leaves = 0
        total_employees = len(original_data)
        
        # Count statuses from the data (assuming status columns are after Employee ID and Name)
        for row in original_data:
            if len(row) > 2:  # Has status columns
                for cell_value in row[2:]:  # Skip Employee ID and Name
                    if isinstance(cell_value, str):
                        if cell_value.strip() == 'Present':
                            total_present += 1
                        elif cell_value.strip() == 'Late':
                            total_late += 1
                        elif cell_value.strip() == 'Absent':
                            total_absent += 1
                        elif cell_value.strip() == 'Leave':
                            total_leaves += 1

        # Insert and style title row
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = "MSN GLOBAL IT SOLUTIONS - Attendance Report"
        title_cell.font = Font(size=16, bold=True, color="FF0000", name="Calibri")
        title_cell.fill = PatternFill("solid", fgColor="000000")  # Black background
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Insert statistics rows
        for i in range(6):  # Increased to 6 for additional stats
            ws.insert_rows(2)
        
        # Add statistics
        stats_data = [
            f"Total Employees: {total_employees}",
            f"Total Present: {total_present}",
            f"Total Late: {total_late}",
            f"Total Absent: {total_absent}",
            f"Total Leaves: {total_leaves}",
            f"Report Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}"
        ]
        
        for i, stat in enumerate(stats_data, 2):
            cell = ws.cell(row=i, column=1)
            cell.value = stat
            cell.font = Font(size=12, bold=True, color="FFFFFF", name="Calibri")
            cell.fill = PatternFill("solid", fgColor="1C1C1C")  # Dark background for stats
            cell.alignment = Alignment(horizontal="left", vertical="center")
            # Merge cells for statistics
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=ws.max_column)

        # Define fills and fonts
        header_fill = PatternFill("solid", fgColor="C80000")  # Red for headers
        row_fill_1 = PatternFill("solid", fgColor="1C1C1C")   # Almost black
        row_fill_2 = PatternFill("solid", fgColor="EDEDED")   # Light gray

        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Style column headers (now at row 9 due to additional stats)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=9, column=col)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF", name="Calibri")  # White bold text
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Style body cells with alternating row colors (starting from row 10)
        for i, row in enumerate(ws.iter_rows(min_row=10, max_row=ws.max_row, min_col=1, max_col=ws.max_column)):
            fill = row_fill_1 if i % 2 == 0 else row_fill_2
            font_color = "FFFFFF" if fill == row_fill_1 else "000000"
            for cell in row:
                cell.fill = fill
                cell.font = Font(name="Calibri", color=font_color)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 3

        # Save styled workbook
        wb.save(file_path)
        
    except Exception as e:
        print(f"Error in style_excel: {e}")

def create_monthly_employee_report():
    """Create monthly employee-wise report with total statistics for each employee"""
    try:
        raw_file = os.path.join(EXCEL_DIR, f'Attendance_Raw_{get_month_year()}.xlsx')
        if os.path.exists(raw_file):
            df = pd.read_excel(raw_file)
            if not df.empty:
                # Load employee data
                employee_df = load_employee_data()
                if employee_df is not None and not employee_df.empty:
                    # Create employee-wise summary
                    employee_summary = []
                    
                    for _, emp_row in employee_df.iterrows():
                        emp_id = emp_row['Employee ID']
                        emp_name = emp_row['Name']
                        
                        # Get all entries for this employee
                        emp_entries = df[df['Employee ID'] == emp_id]
                        
                        # Calculate statistics for this employee
                        total_present = len(emp_entries[emp_entries['Status'] == 'Present'])
                        total_late = len(emp_entries[emp_entries['Status'] == 'Late'])
                        total_absent = len(emp_entries[emp_entries['Status'] == 'Absent'])
                        total_leave = len(emp_entries[emp_entries['Status'] == 'Leave'])
                        total_entries = len(emp_entries)
                        
                        employee_summary.append({
                            'Employee ID': emp_id,
                            'Name': emp_name,
                            'Total Present': total_present,
                            'Total Late': total_late,
                            'Total Absent': total_absent,
                            'Total Leave': total_leave,
                            'Total Entries': total_entries
                        })
                    
                    # Create DataFrame
                    summary_df = pd.DataFrame(employee_summary)
                    
                    # Save to Excel with styling
                    monthly_filename = f'Monthly_Employee_Report_{get_month_year()}.xlsx'
                    create_styled_excel_report(summary_df, monthly_filename)
                    
                    return True, f"✅ Monthly employee report generated: {monthly_filename}"
                else:
                    return False, "❌ No employee data available"
            else:
                return False, "❌ No attendance data available"
        else:
            return False, "❌ No attendance file found"
    except Exception as e:
        return False, f"❌ Error generating monthly report: {str(e)}"

def style_monthly_excel(file_path):
    """Apply professional styling to monthly employee report"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        # Insert and style title row
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"MSN GLOBAL IT SOLUTIONS - Monthly Employee Report ({get_month_year()})"
        title_cell.font = Font(size=16, bold=True, color="FF0000", name="Calibri")
        title_cell.fill = PatternFill("solid", fgColor="000000")  # Black background
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Define fills and fonts
        header_fill = PatternFill("solid", fgColor="C80000")  # Red for headers
        row_fill_1 = PatternFill("solid", fgColor="1C1C1C")   # Almost black
        row_fill_2 = PatternFill("solid", fgColor="EDEDED")   # Light gray

        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Style column headers (now at row 3)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=3, column=col)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF", name="Calibri")  # White bold text
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Style body cells with alternating row colors (starting from row 4)
        for i, row in enumerate(ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column)):
            fill = row_fill_1 if i % 2 == 0 else row_fill_2
            font_color = "FFFFFF" if fill == row_fill_1 else "000000"
            for cell in row:
                cell.fill = fill
                cell.font = Font(name="Calibri", color=font_color)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 3

        # Save styled workbook
        wb.save(file_path)
        
    except Exception as e:
        print(f"Error in style_monthly_excel: {e}")

def create_styled_excel_report(df, filename):
    """Create Excel report with professional styling"""
    try:
        # Save unstyled DataFrame
        file_path = os.path.join(EXCEL_DIR, filename)
        df.to_excel(file_path, index=False, sheet_name="Employee Report")
        
        # Apply styling based on filename
        if 'Monthly_Employee_Report' in filename:
            style_monthly_excel(file_path)
        else:
            style_excel(file_path)
        
        return file_path
    except Exception as e:
        print(f"Error creating styled Excel report: {e}")
        return None

# Login page
def login_page():
    st.markdown("<h1 style='text-align: center; color: #8B0000;'>🔐 MSN GLOBAL IT SOLUTIONS</h1>", unsafe_allow_html=True)
    
    st.markdown("<div class='login-container'>", unsafe_allow_html=True)
    
    with st.form("login_form"):
        username = st.text_input("👤 Username", placeholder="Enter username")
        password = st.text_input("🔒 Password", type="password", placeholder="Enter password")
        
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            submit = st.form_submit_button("🚀 Login", use_container_width=True)
        
        if submit:
            if authenticate(username, password):
                st.session_state.logged_in = True
                st.session_state.username = username
                st.success("✅ Login successful!")
                st.rerun()
            else:
                st.error("❌ Invalid username or password!")
    
    st.markdown("</div>", unsafe_allow_html=True)
    
    st.markdown("### 📋 Login Credentials:")
    st.markdown("""
    - **Username**: `msnglobalit`
    - **Password**: `msnglobalit123`
    """)

# Main dashboard
def main_dashboard():
    # Auto-update daily Excel
    auto_update_daily_excel()
    
    # Continuous face recognition monitoring
    if st.session_state.auto_refresh:
        continuous_face_monitoring()
        monitor_face_recognition_file()
    
    # Smaller aesthetic header
    logo_b64 = get_logo_base64()
    if logo_b64:
        st.markdown("""
        <div style="background-color: #000000; color: white; padding: 1rem; border-radius: 10px; border: 2px solid #DC143C; text-align: center; margin-bottom: 2rem;">
            <div style="display: flex; align-items: center; justify-content: center; gap: 1rem;">
                <img src="data:image/png;base64,{}" width="50" height="50">
                <div>
                    <h1 style="color: #DC143C; margin: 0; font-size: 2rem;">MSN GLOBAL IT SOLUTIONS</h1>
                    <p style="color: white; margin: 0; font-size: 1rem;">AI Attendance Management System</p>
                </div>
            </div>
        </div>
        """.format(logo_b64), unsafe_allow_html=True)
    else:
        st.markdown("""
        <div style="background-color: #000000; color: white; padding: 1rem; border-radius: 10px; border: 2px solid #DC143C; text-align: center; margin-bottom: 2rem;">
            <h1 style="color: #DC143C; margin: 0; font-size: 2rem;">MSN GLOBAL IT SOLUTIONS</h1>
            <p style="color: white; margin: 0; font-size: 1rem;">AI Attendance Management System</p>
        </div>
        """, unsafe_allow_html=True)
    
    # Sidebar with logo and admin panel
    with st.sidebar:
        # Sidebar header with logo
        logo_b64 = get_logo_base64()
        if logo_b64:
            st.markdown("""
            <div style="background-color: #000000; color: white; padding: 1rem; border-radius: 10px; border: 2px solid #DC143C; text-align: center; margin-bottom: 2rem;">
                <img src="data:image/png;base64,{}" width="80" height="80" style="margin-bottom: 1rem;">
                <h3 style="color: #DC143C; margin: 0;">MSN GLOBAL IT</h3>
                <p style="color: white; margin: 0.5rem 0 0 0;">AI Attendance System</p>
            </div>
            """.format(logo_b64), unsafe_allow_html=True)
        else:
            st.markdown("""
            <div style="background-color: #000000; color: white; padding: 1rem; border-radius: 10px; border: 2px solid #DC143C; text-align: center; margin-bottom: 2rem;">
                <h3 style="color: #DC143C; margin: 0;">MSN GLOBAL IT</h3>
                <p style="color: white; margin: 0.5rem 0 0 0;">AI Attendance System</p>
            </div>
            """, unsafe_allow_html=True)
        
        # Admin Panel
        st.markdown("""
        <div style="background-color: #000000; color: white; padding: 1rem; border-radius: 10px; border: 2px solid #DC143C; margin-bottom: 2rem;">
            <h3 style="color: #DC143C; margin: 0;">👤 Admin Panel</h3>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown(f"**👤 User:** {st.session_state.username}")
        st.markdown(f"**📅 Date:** {get_today_date_display()}")
        st.markdown(f"**⏰ Time:** {datetime.now().strftime('%H:%M:%S')}")
        
        # System status
        st.markdown("### ⚙️ System Status")

        # Auto refresh toggle
        auto_refresh_status = st.toggle("🔄 Auto Refresh", value=st.session_state.auto_refresh, key="main_auto_refresh")
        if auto_refresh_status != st.session_state.auto_refresh:
            st.session_state.auto_refresh = auto_refresh_status
            st.rerun()
        
        if st.session_state.auto_refresh:
            st.markdown("**🟢 Auto Refresh:** ON")
        else:
            st.markdown("**🔴 Auto Refresh:** OFF")
        
        # Face recognition status
        json_status = check_json_file_status()
        if json_status["exists"]:
            if json_status["size"] > 0:
                st.markdown("**🟢 Face Recognition:** Data Available")
            else:
                st.markdown("**🟡 Face Recognition:** Waiting for Data")
        else:
            st.markdown("**🔴 Face Recognition:** Not Connected")
        
        # Face Recognition
        st.markdown("### 🔄 Face Recognition")
        
        # Auto-refresh when JSON file changes
        json_status = check_json_file_status()
        if json_status["exists"] and json_status["size"] > 0:
            # Auto-process face recognition data
            process_multiple_face_recognition()
        
        # Manual refresh button
        if st.button("🔄 REFRESH ATTENDANCE", type="primary", use_container_width=True, key="main_refresh_button"):
            process_multiple_face_recognition()
        

        

        
        # CSV upload
        st.markdown("### 📁 Upload Employee CSV")
        
        # Show current uploaded file if exists
        if st.session_state.uploaded_csv_data is not None:
            st.success(f"✅ Current file: {st.session_state.uploaded_csv_filename}")
            col1, col2 = st.columns([3, 1])
            with col1:
                if st.button("🗑️ Remove Current File", type="secondary", use_container_width=True):
                    st.session_state.uploaded_csv_data = None
                    st.session_state.uploaded_csv_filename = None
                    st.success("✅ File removed!")
                    st.rerun()
            with col2:
                if st.button("📊 View Data", type="secondary", use_container_width=True):
                    df_display = pd.DataFrame(st.session_state.uploaded_csv_data)
                    st.dataframe(df_display, use_container_width=True)
        
        uploaded_file = st.file_uploader("Choose CSV file", type=['csv'])
        if uploaded_file is not None:
            try:
                df_upload = pd.read_csv(uploaded_file)
                if 'Employee ID' in df_upload.columns and 'Name' in df_upload.columns:
                    emp_file = os.path.join(os.path.dirname(BASE_DIR), 'face_recognition_app', 'shared', 'employees_data.csv')
                    df_upload.to_csv(emp_file, index=False)
                    st.session_state.uploaded_csv_data = df_upload.to_dict('records') # Store data in session state
                    st.session_state.uploaded_csv_filename = uploaded_file.name
                    st.success("✅ Employee data updated!")
                    st.rerun()
                else:
                    st.error("❌ CSV must have 'Employee ID' and 'Name' columns")
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
        
        # Logout
        st.markdown("### 🚪 Account")
        if st.button("🚪 Logout", type="secondary", use_container_width=True):
            st.session_state.logged_in = False
            st.session_state.username = None
            st.rerun()
    
    # Main content
    col1, col2 = st.columns([4, 1])
    
    with col1:
        # Quick controls
        st.markdown("### 🎛️ Quick Controls")
        control_col1, control_col2, control_col3, control_col4 = st.columns(4)
        
        with control_col1:
            if st.button("📊 Generate Report", type="primary", use_container_width=True, key="gen_report"):
                raw_file = os.path.join(EXCEL_DIR, f'Attendance_Raw_{get_month_year()}.xlsx')
                if os.path.exists(raw_file):
                    try:
                        df = pd.read_excel(raw_file)
                        if not df.empty:
                            actual_filename = generate_wide_excel(df, f'Attendance_Report_{get_month_year()}.xlsx')
                            st.success("✅ Excel report generated and saved!")
                        else:
                            st.markdown("""
                            <div style="background-color: #FF9800; color: white; padding: 10px; border-radius: 8px; margin: 5px 0; border: 2px solid #F57C00;">
                                <p style="margin: 0; color: white;">⚠ No data available for report</p>
                            </div>
                            """, unsafe_allow_html=True)
                    except Exception as e:
                        st.error(f"❌ Error generating report: {str(e)}")
                else:
                    st.markdown("""
                    <div style="background-color: #FF9800; color: white; padding: 10px; border-radius: 8px; margin: 5px 0; border: 2px solid #F57C00;">
                        <p style="margin: 0; color: white;">⚠ No attendance data available</p>
                    </div>
                    """, unsafe_allow_html=True)

        with control_col2:
            if st.button("📈 Monthly Report", type="secondary", use_container_width=True, key="monthly_report"):
                success, message = create_monthly_employee_report()
                if success:
                    st.success(message)
                else:
                    st.error(message)
        
        with control_col3:
            if st.button("🗑️ Clear Entries", type="secondary", use_container_width=True, key="clear_entries"):
                raw_file = os.path.join(EXCEL_DIR, f'Attendance_Raw_{get_month_year()}.xlsx')
                report_file = os.path.join(EXCEL_DIR, f'Attendance_Report_{get_month_year()}.xlsx')
                monthly_file = os.path.join(EXCEL_DIR, f'Monthly_Employee_Report_{get_month_year()}.xlsx')
                realtime_file = os.path.join(EXCEL_DIR, 'realtime_attendance.xlsx')
                deleted_files = []
                # Remove today's entries from raw file
                if os.path.exists(raw_file):
                    try:
                        df = pd.read_excel(raw_file)
                        today_date = get_today_date()
                        # Remove today's entries
                        df_filtered = df[df['Date'] != today_date]
                        df_filtered.to_excel(raw_file, index=False)
                        deleted_files.append(raw_file)
                    except Exception as e:
                        st.error(f"❌ Error clearing entries: {str(e)}")
                # Delete report files if they exist
                for f in [report_file, monthly_file, realtime_file]:
                    if os.path.exists(f):
                        try:
                            os.remove(f)
                            deleted_files.append(f)
                        except Exception as e:
                            st.error(f"❌ Error deleting file {f}: {str(e)}")
                if deleted_files:
                    st.success(f"✅ Cleared today's entries and deleted files: {', '.join([os.path.basename(f) for f in deleted_files])}")
                    st.rerun()
                else:
                    st.markdown("""
                    <div style="background-color: #2196F3; color: white; padding: 10px; border-radius: 8px; margin: 5px 0; border: 2px solid #1976D2;">
                        <p style="margin: 0; color: white;">ℹ️ No data to clear</p>
                    </div>
                    """, unsafe_allow_html=True)
        
        with control_col4:
            if st.button("📊 Export Data", type="secondary", use_container_width=True, key="export_data"):
                raw_file = os.path.join(EXCEL_DIR, f'Attendance_Raw_{get_month_year()}.xlsx')
                if os.path.exists(raw_file):
                    try:
                        df = pd.read_excel(raw_file)
                        today_df = df[df['Date'] == get_today_date()]
                        if not today_df.empty:
                            csv_data = today_df.to_csv(index=False)
                            # Use a container to show download button
                            download_container = st.container()
                            with download_container:
                                st.download_button(
                                    label="📥 Download Today's CSV",
                                    data=csv_data,
                                    file_name=f"attendance_{get_today_date().replace('/', '_')}.csv",
                                    mime="text/csv",
                                    key="download_csv"
                                )
                        else:
                            st.markdown("""
                            <div style="background-color: #2196F3; color: white; padding: 10px; border-radius: 8px; margin: 5px 0; border: 2px solid #1976D2;">
                                <p style="margin: 0; color: white;">ℹ️ No data to export today</p>
                            </div>
                            """, unsafe_allow_html=True)
                    except Exception as e:
                        st.error(f"❌ Error exporting data: {str(e)}")
                else:
                    st.markdown("""
                    <div style="background-color: #2196F3; color: white; padding: 10px; border-radius: 8px; margin: 5px 0; border: 2px solid #1976D2;">
                        <p style="margin: 0; color: white;">ℹ️ No data available for export</p>
                    </div>
                    """, unsafe_allow_html=True)
        
        # Status metrics
        raw_file = os.path.join(EXCEL_DIR, f'Attendance_Raw_{get_month_year()}.xlsx')
        if os.path.exists(raw_file):
            try:
                df = pd.read_excel(raw_file)
                today_count = len(df[df['Date'] == get_today_date()])
                
                # Calculate metrics for today
                today_df = df[df['Date'] == get_today_date()]
                present_count = len(today_df[today_df['Status'] == 'Present'])
                late_count = len(today_df[today_df['Status'] == 'Late'])
                absent_count = len(today_df[today_df['Status'] == 'Absent'])
                
                # Calculate total metrics (all time)
                total_present = len(df[df['Status'] == 'Present'])
                total_late = len(df[df['Status'] == 'Late'])
                total_absent = len(df[df['Status'] == 'Absent'])
                
                # Calculate leave (assuming leave is marked as 'Leave' or 'Absent' for specific cases)
                total_leave = len(df[df['Status'] == 'Leave']) if 'Leave' in df['Status'].values else 0
                
                # Display today's metrics
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("✅ Present Today", present_count)
                with col2:
                    st.metric("⏰ Late Today", late_count)
                with col3:
                    st.metric("❌ Absent Today", absent_count)
                with col4:
                    st.metric("📊 Total Entries", today_count)
                
            except Exception as e:
                st.error(f"❌ Error loading metrics: {str(e)}")
        else:
            # Show zero metrics if no data
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("✅ Present Today", 0)
            with col2:
                st.metric("⏰ Late Today", 0)
            with col3:
                st.metric("❌ Absent Today", 0)
            with col4:
                st.metric("📊 Total Entries", 0)
        
        # Manual attendance
        st.markdown("### 👤 Manual Attendance")
        
        employee_df = load_employee_data()
        if employee_df is not None and not employee_df.empty:
            employee_options = [f"{row['Employee ID']} - {row['Name']}" for _, row in employee_df.iterrows()]
            selected_employee = st.selectbox("Select Employee", employee_options)
            
            if st.button("✅ Mark Attendance", type="primary"):
                if selected_employee:
                    emp_id = selected_employee.split(" - ")[0]
                    name = selected_employee.split(" - ")[1]
                    success, message = mark_attendance(emp_id, name)
                    if success:
                        st.success(message)
                        st.rerun()
                    else:
                        st.error(message)
        else:
            st.markdown("""
            <div style="background-color: #FF9800; color: white; padding: 10px; border-radius: 8px; margin: 5px 0; border: 2px solid #F57C00;">
                <p style="margin: 0; color: white;">⚠️ Please upload employee data first</p>
            </div>
            """, unsafe_allow_html=True)
        
        # Current time info
        st.markdown("### ⏰ Current Time")
        current_time = datetime.now()
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("🕐 Current Time", current_time.strftime("%H:%M:%S"))
        with col2:
            st.metric("📅 Date", current_time.strftime("%Y-%m-%d"))
        with col3:
            st.metric("📊 Status", "Active")
        
        # Attendance data with enhanced display
        st.markdown("### 📊 Today's Attendance")
        
        raw_file = os.path.join(EXCEL_DIR, f'Attendance_Raw_{get_month_year()}.xlsx')
        if os.path.exists(raw_file):
            try:
                df = pd.read_excel(raw_file)
                today_df = df[df['Date'] == get_today_date()]
                
                if not today_df.empty:
                    # Show enhanced information including time
                    if 'Entry_Time' in today_df.columns:
                        # Format time for display - use .copy() to avoid SettingWithCopyWarning
                        display_df = today_df[['Employee ID', 'Name', 'Status', 'Entry_Time']].copy()
                        display_df['Entry_Time_Display'] = display_df['Entry_Time'].apply(
                            lambda x: x.strftime('%H:%M') if hasattr(x, 'strftime') else str(x)
                        )
                        display_df = display_df[['Employee ID', 'Name', 'Status', 'Entry_Time_Display']]
                        display_df.columns = ['Employee ID', 'Name', 'Status', 'Entry Time']
                    else:
                        display_df = today_df[['Employee ID', 'Name', 'Status']].copy()
                    
                    # Add color coding to the dataframe
                    st.dataframe(display_df, use_container_width=True)
                    
                    # Show summary statistics
                    st.markdown("#### 📈 Attendance Summary")
                    status_counts = today_df['Status'].value_counts()
                    for status, count in status_counts.items():
                        if status == 'Present':
                            st.success(f"✅ {status}: {count} employee(s)")
                        elif status == 'Late':
                            st.warning(f"⏰ {status}: {count} employee(s)")
                        elif status == 'Absent':
                            st.error(f"❌ {status}: {count} employee(s)")
                        else:
                            st.info(f"📊 {status}: {count} employee(s)")
                else:
                    st.markdown("""
                    <div style="background-color: #2196F3; color: white; padding: 10px; border-radius: 8px; margin: 5px 0; border: 2px solid #1976D2;">
                        <p style="margin: 0; color: white;">ℹ️ No attendance data for today</p>
                    </div>
                    """, unsafe_allow_html=True)
            except Exception as e:
                st.error(f"❌ Error loading attendance data: {str(e)}")
        else:
            st.markdown("""
            <div style="background-color: #2196F3; color: white; padding: 10px; border-radius: 8px; margin: 5px 0; border: 2px solid #1976D2;">
                <p style="margin: 0; color: white;">ℹ️ No attendance data available</p>
            </div>
            """, unsafe_allow_html=True)
    
    with col2:
        # Empty column for balance
        pass
    
    # Auto refresh
    if st.session_state.auto_refresh:
        current_time = time.time()
        if current_time - st.session_state.refresh_timestamp > 5:
            st.session_state.refresh_timestamp = current_time
            st.cache_data.clear()
            st.rerun()

def continuous_face_monitoring():
    """Continuously monitor for face recognition data and auto-save"""
    try:
        current_time = time.time()
        # Check every 2 seconds for better performance
        if current_time - st.session_state.last_face_check > 2.0:
            st.session_state.last_face_check = current_time
            
            # Check for face recognition data
            success, message = process_multiple_face_recognition()
            if success:
                # Store notification in session state
                st.session_state.last_notification = message
                st.session_state.notification_type = "success"
                time.sleep(0.1)  # Brief pause to show the message
                st.rerun()
    except Exception as e:
        print(f"Error in continuous face monitoring: {e}")

def monitor_face_recognition_file():
    """Monitor the face recognition JSON file for changes and auto-refresh"""
    try:
        if os.path.exists(RECOGNIZED_ID_FILE):
            # Check if file has been modified recently
            file_stat = os.stat(RECOGNIZED_ID_FILE)
            current_time = time.time()
            
            # If file was modified in the last 3 seconds and we haven't checked recently
            if (current_time - file_stat.st_mtime < 3 and 
                current_time - st.session_state.last_face_check > 2.0):
                st.session_state.last_face_check = current_time
                success, message = process_multiple_face_recognition()
                if success:
                    # Store notification in session state
                    st.session_state.last_notification = message
                    st.session_state.notification_type = "success"
                    time.sleep(0.1)
                    st.rerun()
        else:
            # Keep monitoring even if no JSON file exists
            current_time = time.time()
            if current_time - st.session_state.last_face_check > 5:  # Check every 5 seconds
                st.session_state.last_face_check = current_time
                # Keep the system ready for face detection
    except Exception as e:
        print(f"Error monitoring face recognition file: {e}")

def process_multiple_face_recognition():
    """Process multiple face recognition IDs from JSON file"""
    try:
        if not os.path.exists(RECOGNIZED_ID_FILE):
            return False, "No face recognition data available"
        
        file_size = os.path.getsize(RECOGNIZED_ID_FILE)
        
        if file_size == 0:
            return False, "No new detection data"
        
        # Read JSON content
        try:
            with open(RECOGNIZED_ID_FILE, 'r') as f:
                data = json.load(f)
        except Exception as e:
            return False, f"Error reading JSON file: {str(e)}"
        
        # Handle both single entry and multiple entries
        entries = []
        if isinstance(data, list):
            entries = data  # Multiple entries
        else:
            entries = [data]  # Single entry
        
        if not entries:
            return False, "No valid entries found"
        
        # Process all entries (removed time filtering for now)
        success_count = 0
        messages = []
        processed_ids = set()  # Track processed IDs to avoid duplicates
        
        for i, entry in enumerate(entries):
            emp_id = entry.get('employee_id')
            name = entry.get('name', '')
            unique_id = entry.get('unique_id', '')
            
            if not emp_id:
                continue
            
            if not unique_id:
                continue
            
            # Skip if already processed this unique ID
            if unique_id in processed_ids:
                continue
            
            processed_ids.add(unique_id)
            
            # Check for duplicate before processing
            if is_duplicate_attendance_streamlit(emp_id):
                messages.append(f"⚠️ {emp_id}: Already marked today")
                continue
            
            if not name:
                # Try to get name from employee CSV data
                try:
                    employee_df = load_employee_data()
                    if employee_df is not None and not employee_df.empty:
                        emp_row = employee_df[employee_df['Employee ID'] == emp_id]
                        if not emp_row.empty:
                            name = emp_row['Name'].iloc[0]
                except Exception as e:
                    name = f"Employee {emp_id}"
            
            success, message = mark_attendance(emp_id, name or f"Employee {emp_id}")
            
            if success:
                success_count += 1
                messages.append(f"✅ {name or emp_id} ({emp_id})")
            else:
                messages.append(f"❌ {emp_id}: {message}")
        
        # Clear the JSON file after processing all entries
        try:
            os.remove(RECOGNIZED_ID_FILE)
        except Exception as e:
            print(f"⚠️ Error clearing JSON file: {e}")
        
        if success_count > 0:
            result_message = f"✅ Face Recognition: {success_count} attendance(s) marked - {', '.join(messages)}"
            print(f"🎉 Success: {result_message}")
            return True, result_message
        else:
            result_message = f"No new attendance marked - {', '.join(messages)}"
            print(f"ℹ️ No success: {result_message}")
            return False, result_message
            
    except json.JSONDecodeError as e:
        print(f"❌ JSON decode error: {e}")
        # Clear corrupted JSON file
        try:
            os.remove(RECOGNIZED_ID_FILE)
            print("🗑️ Cleared corrupted JSON file")
        except:
            pass
        return False, "Cleared corrupted data"
    except Exception as e:
        print(f"❌ Unexpected error: {e}")
        return False, f"Error processing data: {str(e)}"
    
    return False, None

def process_multiple_face_recognition_enhanced():
    """Enhanced version with detailed feedback and better duplicate prevention"""
    try:
        if not os.path.exists(RECOGNIZED_ID_FILE):
            return False, "ℹ️ No JSON file found - Face recognition app may not be running", None
        
        file_size = os.path.getsize(RECOGNIZED_ID_FILE)
        if file_size == 0:
            return False, "ℹ️ JSON file is empty - No new detection data", None
        
        # Read JSON content
        with open(RECOGNIZED_ID_FILE, 'r') as f:
            data = json.load(f)
        
        # Handle both single entry and multiple entries
        entries = []
        if isinstance(data, list):
            entries = data  # Multiple entries
        else:
            entries = [data]  # Single entry
        
        if not entries:
            return False, "❌ No valid entries found in JSON", None
        
        # Filter entries to only process recent ones (within last 30 seconds)
        current_time = datetime.now()
        recent_entries = []
        for entry in entries:
            try:
                timestamp_str = entry.get('timestamp', '')
                if timestamp_str:
                    entry_time = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
                    time_diff = (current_time - entry_time).total_seconds()
                    if time_diff < 30:  # Only process entries from last 30 seconds
                        recent_entries.append(entry)
                    else:
                        print(f"⏰ Skipping old entry for {entry.get('employee_id', 'Unknown')} (age: {time_diff:.1f}s)")
            except Exception as e:
                print(f"❌ Error parsing timestamp: {e}")
                continue
        
        if not recent_entries:
            return False, "❌ No recent entries found in JSON", None
        
        entries = recent_entries  # Use only recent entries
        
        # Processing details for feedback
        processing_details = {
            "total_entries": len(entries),
            "processed_ids": set(),
            "duplicate_entries": [],
            "successful_entries": [],
            "failed_entries": [],
            "skipped_entries": []
        }
        
        success_count = 0
        messages = []
        
        for i, entry in enumerate(entries):
            emp_id = entry.get('employee_id')
            name = entry.get('name', '')
            unique_id = entry.get('unique_id', '')
            timestamp = entry.get('timestamp', '')
            
            # Validate entry
            if not emp_id or not unique_id:
                processing_details["skipped_entries"].append(f"Entry {i+1}: Missing employee_id or unique_id")
                continue
            
            # Check for duplicate unique_id
            if unique_id in processing_details["processed_ids"]:
                processing_details["duplicate_entries"].append(f"{emp_id} (unique_id: {unique_id})")
                continue
            
            processing_details["processed_ids"].add(unique_id)
            
            # Check for duplicate attendance (same employee, same day)
            if is_duplicate_attendance_streamlit(emp_id):
                processing_details["skipped_entries"].append(f"{emp_id}: Already marked attendance today")
                messages.append(f"⚠️ {emp_id}: Already marked today")
                continue
            
            # Get employee name if not provided
            if not name:
                try:
                    employee_df = load_employee_data()
                    if employee_df is not None and not employee_df.empty:
                        emp_row = employee_df[employee_df['Employee ID'] == emp_id]
                        if not emp_row.empty:
                            name = emp_row['Name'].iloc[0]
                except Exception as e:
                    print(f"Error getting employee name from CSV: {e}")
                    name = f"Employee {emp_id}"
            
            # Mark attendance
            success, message = mark_attendance(emp_id, name or f"Employee {emp_id}")
            if success:
                success_count += 1
                processing_details["successful_entries"].append({
                    "employee_id": emp_id,
                    "name": name or f"Employee {emp_id}",
                    "timestamp": timestamp,
                    "unique_id": unique_id
                })
                messages.append(f"✅ {name or emp_id} ({emp_id})")
            else:
                processing_details["failed_entries"].append({
                    "employee_id": emp_id,
                    "error": message
                })
                messages.append(f"❌ {emp_id}: {message}")
        
        # Clear the JSON file after processing
        try:
            os.remove(RECOGNIZED_ID_FILE)
        except:
            pass
        
        # Prepare detailed feedback
        details = {
            "📊 Processing Summary": {
                "Total entries found": processing_details["total_entries"],
                "Unique entries processed": len(processing_details["processed_ids"]),
                "Successful attendance marks": len(processing_details["successful_entries"]),
                "Failed attempts": len(processing_details["failed_entries"]),
                "Skipped entries": len(processing_details["skipped_entries"]),
                "Duplicate entries": len(processing_details["duplicate_entries"])
            }
        }
        
        if processing_details["successful_entries"]:
            details["✅ Successful Entries"] = processing_details["successful_entries"]
        
        if processing_details["failed_entries"]:
            details["❌ Failed Entries"] = processing_details["failed_entries"]
        
        if processing_details["skipped_entries"]:
            details["⚠️ Skipped Entries"] = processing_details["skipped_entries"]
        
        if processing_details["duplicate_entries"]:
            details["🔄 Duplicate Entries"] = processing_details["duplicate_entries"]
        
        if success_count > 0:
            return True, f"✅ Face Recognition: {success_count} attendance(s) marked successfully", details
        else:
            return False, f"ℹ️ No new attendance marked - {', '.join(messages)}", details
            
    except json.JSONDecodeError:
        # Clear corrupted JSON file
        try:
            os.remove(RECOGNIZED_ID_FILE)
        except:
            pass
        return False, "❌ Corrupted JSON data - File cleared", {"❌ Error": "JSON file was corrupted and has been cleared"}
    except Exception as e:
        print(f"Error reading face recognition JSON: {e}")
        return False, f"❌ Error processing JSON: {str(e)}", {"❌ Error": str(e)}

def check_json_file_status():
    """Check the status of the JSON file and return detailed information"""
    try:
        if not os.path.exists(RECOGNIZED_ID_FILE):
            return {
                "exists": False,
                "size": 0,
                "content": None,
                "status": "File not found"
            }
        
        file_size = os.path.getsize(RECOGNIZED_ID_FILE)
        if file_size == 0:
            return {
                "exists": True,
                "size": 0,
                "content": None,
                "status": "File is empty"
            }
        
        # Try to read the JSON content
        try:
            with open(RECOGNIZED_ID_FILE, 'r') as f:
                content = json.load(f)
            return {
                "exists": True,
                "size": file_size,
                "content": content,
                "status": "File contains valid JSON"
            }
        except json.JSONDecodeError:
            return {
                "exists": True,
                "size": file_size,
                "content": None,
                "status": "File contains invalid JSON"
            }
        except Exception as e:
            return {
                "exists": True,
                "size": file_size,
                "content": None,
                "status": f"Error reading file: {str(e)}"
            }
            
    except Exception as e:
        return {
            "exists": False,
            "size": 0,
            "content": None,
            "status": f"Error checking file: {str(e)}"
        }

def is_duplicate_attendance_streamlit(emp_id):
    """Enhanced duplicate check - Check if this specific employee already marked attendance today"""
    today_date = get_today_date()
    
    # Check multiple data sources for duplicates
    sources_checked = []
    
    # 1. Check Excel file (primary source)
    try:
        raw_file = os.path.join(EXCEL_DIR, f'Attendance_Raw_{get_month_year()}.xlsx')
        if os.path.exists(raw_file):
            df = pd.read_excel(raw_file)
            if not df.empty and 'Date' in df.columns and 'Employee ID' in df.columns:
                today_entries = df[df['Date'] == today_date]
                this_employee_entries = today_entries[today_entries['Employee ID'] == emp_id]
                if len(this_employee_entries) > 0:
                    sources_checked.append(f"Excel: {len(this_employee_entries)} entry(ies) found")
                    return True
                else:
                    sources_checked.append("Excel: No entries found")
            else:
                sources_checked.append("Excel: Invalid data structure")
    except Exception as e:
        sources_checked.append(f"Excel: Error - {str(e)}")
    
    # 2. Check CSV file (backup source)
    try:
        csv_file = os.path.join(os.path.dirname(BASE_DIR), 'face_recognition_app', 'shared', 'attendance_log.csv')
        if os.path.exists(csv_file):
            df = pd.read_csv(csv_file, on_bad_lines='skip', engine='python')
            if not df.empty and 'Date' in df.columns and 'Employee ID' in df.columns:
                today_entries = df[df['Date'] == today_date]
                this_employee_entries = today_entries[today_entries['Employee ID'] == emp_id]
                if len(this_employee_entries) > 0:
                    sources_checked.append(f"CSV: {len(this_employee_entries)} entry(ies) found")
                    return True
                else:
                    sources_checked.append("CSV: No entries found")
            else:
                sources_checked.append("CSV: Invalid data structure")
    except Exception as e:
        sources_checked.append(f"CSV: Error - {str(e)}")
    
    # 3. Check for any other attendance files in the system
    try:
        # Check if there are any other attendance files for today
        attendance_files = []
        for file in os.listdir(EXCEL_DIR):
            if file.startswith('Attendance_') and file.endswith('.xlsx'):
                attendance_files.append(file)
        
        for file in attendance_files:
            try:
                file_path = os.path.join(EXCEL_DIR, file)
                df = pd.read_excel(file_path)
                if not df.empty and 'Date' in df.columns and 'Employee ID' in df.columns:
                    today_entries = df[df['Date'] == today_date]
                    this_employee_entries = today_entries[today_entries['Employee ID'] == emp_id]
                    if len(this_employee_entries) > 0:
                        sources_checked.append(f"{file}: {len(this_employee_entries)} entry(ies) found")
                        return True
            except Exception as e:
                sources_checked.append(f"{file}: Error - {str(e)}")
    except Exception as e:
        sources_checked.append(f"File scan: Error - {str(e)}")
    
    return False  # This employee hasn't marked attendance today

def get_last_sync_time():
    """Get the last sync time from the JSON file"""
    try:
        if os.path.exists(RECOGNIZED_ID_FILE):
            file_stat = os.stat(RECOGNIZED_ID_FILE)
            last_modified = datetime.fromtimestamp(file_stat.st_mtime)
            return last_modified.strftime('%H:%M:%S')
        return "Never"
    except Exception as e:
        return "Unknown"

def get_sync_status():
    """Get detailed sync status"""
    try:
        if os.path.exists(RECOGNIZED_ID_FILE):
            file_stat = os.stat(RECOGNIZED_ID_FILE)
            current_time = time.time()
            
            if current_time - file_stat.st_mtime < 60:  # Within last minute
                return "🟢 Active - Recently updated"
            elif current_time - file_stat.st_mtime < 300:  # Within last 5 minutes
                return "🟡 Active - Updated recently"
            else:
                return "🟡 Active - Ready for detection"
        else:
            return "🟢 Ready - Waiting for face detection"
    except Exception as e:
        return "🟢 Ready - System active"

def get_today_attendance_summary():
    """Get summary of today's attendance"""
    today_date = get_today_date()
    summary = {'present': 0, 'late': 0, 'absent': 0, 'total': 0}
    
    try:
        raw_file = os.path.join(EXCEL_DIR, f'Attendance_Raw_{get_month_year()}.xlsx')
        if os.path.exists(raw_file):
            df = pd.read_excel(raw_file)
            if not df.empty and 'Date' in df.columns and 'Status' in df.columns:
                today_entries = df[df['Date'] == today_date]
                if not today_entries.empty:
                    summary['present'] = len(today_entries[today_entries['Status'] == 'Present'])
                    summary['late'] = len(today_entries[today_entries['Status'] == 'Late'])
                    summary['absent'] = len(today_entries[today_entries['Status'] == 'Absent'])
                    summary['total'] = len(today_entries)
                    return summary
    except Exception as e:
        print(f"Error getting attendance summary: {e}")
    
    return None

# Main app logic
def main():
    if not st.session_state.logged_in:
        login_page()
    else:
        main_dashboard()

if __name__ == "__main__":
    main() 